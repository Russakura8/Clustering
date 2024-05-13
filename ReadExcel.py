import openpyxl



class ReadExcel:
    def __init__(self):
        self.colums = dict()
        self.students = []
        self.text_to_numbers = dict()
        self.text_to_norm_numbers = dict()
        self.mas_of_data = []
        self.mas_of_norm_data = []
        self.not_interested = ['Отметка времени',
                               'Укажите, как преподаватели оценивают результаты самостоятельной работы (если в предыдущем вопросе вы выбрали ответ "да")',
                  '14 (а) Укажите e-mail тьютора, если помните',
                               '14 (б) Укажите номер телефона тьютора, если помните',
                               '14 (в) Укажите страничку в социальной сети (ВКонтакте, Фейсбук, Твиттер, Одноклассники и др.) тьютора, если помните',
                  '14 (г) Укажите скайп тьютора, если помните',
                               '14 (д) Укажите другой способ контакта с тьютором, если помните',
                  '15. Перечислите мероприятия, которые были организованы тьютором для Вашей студенческой группы в ПРОШЛОМ семестре',
                  '5. Какие инновационные методы обучения используют преподаватели на занятиях?',
                  'Поставьте флажок напротив тех вопросов, о которых вам известно, что по ним в ТвГУ ведется профилактика:',
                  "2. Почему Вы выбрали для обучения именно ТвГУ?",
                  '4. Какие проблемы Вы видите в организации учебного процесса? (выберите, пожалуйста, не более 3-х вариантов)',
                  '8. Какие проблемы возникают у Вас в процессе самостоятельной работы?',
                  'Укажите свои ФИО (по желанию)',
                               '25. Ваши пожелания по повышению качества питания в столовых и буфетах ТвГУ?',
                               '22. Опишите Ваше питание в ТвГУ',
                               '18. Если обращались, вы удовлетворены оказанной Вам помощью?',
                               '30. Какими услугами в Центре содействия трудоустройству молодежи ТвГУ или Молодежной бирже труда ТвГУ Вы уже воспользовались? (можно указать несколько вариантов)']


    def read(self):
        workbook = openpyxl.load_workbook('Удовлетворенность студентов УВР в вузе, весна 2022 (Ответы).xlsx')
        sheet = workbook.active
        self.colums['Выберите ваше направление/специальность'] = set()
        for row in sheet.iter_rows(values_only=False):
            temp = dict()
            for cell in row:
                global_col = sheet[cell.column_letter + '1'].value
                if (global_col not in self.colums) and (global_col != 'Выберите ваше направление') and \
                        (global_col != 'Выберите Ваш факультет/направление') and global_col not in self.not_interested:
                    self.colums[global_col] = set()
                if cell.value is not None and global_col != cell.value:
                    if global_col == 'Выберите ваше направление':
                        temp['Выберите ваше направление/специальность'] = cell.value
                        self.colums['Выберите ваше направление/специальность'].add(cell.value)
                    elif global_col == 'Выберите Ваш факультет/направление':
                        temp['Выберите Ваш факультет'] = cell.value.split(' /')[0]
                        temp['Выберите ваше направление/специальность'] = cell.value.split(' /')[1]
                        self.colums['Выберите Ваш факультет'].add(cell.value.split(' /')[0])
                        self.colums['Выберите ваше направление/специальность'].add(cell.value.split(' /')[1])
                    elif global_col not in self.not_interested:
                        temp[global_col] = cell.value
                        self.colums[global_col].add(cell.value)

            self.students.append(temp)

        self.students.pop(0)
        return self.students

    def convertTextToNumbers(self):
        for column, value in self.colums.items():
            count = 0
            temp = dict()
            for el in value:
                temp[el] = count
                count += 1
            self.text_to_numbers[column] = temp.copy()

            for el in value:
                temp[el] /= (count - 1)

            self.text_to_norm_numbers[column] = temp.copy()

        print(self.text_to_norm_numbers)
        print(self.text_to_numbers)

    def do_mas(self):
        workbook = openpyxl.load_workbook('Удовлетворенность студентов УВР в вузе, весна 2022 (Ответы).xlsx')
        sheet = workbook.active

        for row in sheet.iter_rows(values_only=False):
            temp = []
            temp2 = []
            for cell in row:
                global_col = sheet[cell.column_letter + '1'].value
                if cell.value is not None and global_col != cell.value:
                    if global_col == 'Выберите ваше направление':
                        temp.append(self.text_to_numbers['Выберите ваше направление/специальность'][cell.value])
                        temp2.append(self.text_to_norm_numbers['Выберите ваше направление/специальность'][cell.value])

                    elif global_col == 'Выберите Ваш факультет/направление':
                        temp.append(self.text_to_numbers['Выберите Ваш факультет'][cell.value.split(' /')[0]])
                        temp.append(self.text_to_numbers['Выберите ваше направление/специальность'][cell.value.split(' /')[1]])
                        temp2.append(self.text_to_norm_numbers['Выберите Ваш факультет'][cell.value.split(' /')[0]])
                        temp2.append(self.text_to_norm_numbers['Выберите ваше направление/специальность'][cell.value.split(' /')[1]])


                    elif global_col not in self.not_interested:
                        temp.append(self.text_to_numbers[global_col][cell.value])
                        temp2.append(self.text_to_norm_numbers[global_col][cell.value])

            self.mas_of_data.append(temp)
            self.mas_of_norm_data.append(temp2)

        self.mas_of_data.pop(0)
        self.mas_of_norm_data.pop(0)
        with open('normdata.txt', 'w') as f:
            f.write(str(self.mas_of_norm_data))

        with open('students.txt', 'w') as f:
            f.write(str(self.students))





cl = ReadExcel()
cl.read()
cl.convertTextToNumbers()
cl.do_mas()

