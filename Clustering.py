import numpy as np
from openpyxl import Workbook
from sklearn import cluster, mixture
import matplotlib.pyplot as plt
from sklearn.decomposition import PCA
import openpyxl

with open('normdata.txt', 'r') as f:
    _data = f.read()

data = eval(_data)


with open('students.txt', 'r') as f:
    _data = f.read()

students = eval(_data)


data = np.array(data)
np.set_printoptions(threshold=np.inf)




dbscan = cluster.DBSCAN(min_samples=3, eps=2.16).fit(data)
spectral = cluster.SpectralClustering(n_clusters=6).fit(data)
kmeans = cluster.KMeans(n_clusters=6).fit(data)
gmm = mixture.GaussianMixture(n_components = 6).fit(data)
birch = cluster.Birch().fit(data)


methods = ["dbscan", "spectral", "kmeans", "gmm", "birch"]

pca = PCA(n_components=2)
pca.fit(data)
new_data = pca.transform(data)

for method in methods:
    if method == "gmm":
        labels = eval(f"{method}.predict(data)")
    else:
        labels = eval(f"{method}.labels_")


    plt.scatter(new_data[:, 0], new_data[:, 1], c=labels, s=50, cmap='viridis')
    plt.savefig(method + ".png")


    wb = Workbook()


    for i in range(max(labels) + 1):
        ws = wb.create_sheet(str(i))
        count = 1
        for key in students[0]:
            cell = ws.cell(row=1, column=count)
            count += 1
            cell.value = key
        count = 2
        for j in range(len(labels)):
            if labels[j] == i:
                count2 = 1
                for key, value in students[j].items():
                    cell =  ws.cell(row=count, column=count2)
                    cell.value = value
                    count2 += 1
                count += 1

    page  = wb['Sheet']
    wb.remove(page)

    wb.save(method + ".xlsx")
    wb.close()

